"""Carrega e persiste o checklist Anexo IV (JSON embutido ou Excel)."""
from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from inspections.models import ChecklistItem, ChecklistSection, ChecklistVersion

SECTION_RE = re.compile(r"^(\d+)\.\s+(.+)$")
BUNDLED_CHECKLIST_JSON = Path(__file__).resolve().parent.parent / "data" / "checklist_anexo_iv.json"


def parse_checklist_xlsx(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    start_idx = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any("PREENCHER COM NÚMERO" in c for c in cells):
            start_idx = i + 1
            break
    rows = rows[start_idx:]

    sections: list[dict] = []
    current: dict | None = None
    item_order = 0
    seen_titles: set[str] = set()

    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        line = next((c for c in cells if c), "")

        if line in ("Total", "Total Geral") or line in ("C", "NC", "NA", "Descrição"):
            continue

        sec_match = SECTION_RE.match(line)
        if sec_match and len(line) < 120 and "?" not in line:
            section_num = int(sec_match.group(1))
            title = sec_match.group(2).strip()
            if title.upper().startswith("TOTAL"):
                continue
            full_title = f"{section_num}. {title}"
            if full_title in seen_titles:
                current = next((s for s in sections if s["title"] == full_title), None)
                item_order = len(current["items"]) if current else 0
                continue
            seen_titles.add(full_title)
            current = {"order": section_num, "title": full_title, "items": []}
            sections.append(current)
            item_order = 0
            continue

        if line.upper().startswith("OUTRAS SITUAÇÕES"):
            if line not in seen_titles:
                seen_titles.add(line)
                current = {"order": 21, "title": line, "items": []}
                sections.append(current)
                item_order = 0
            continue

        if current and "?" in line:
            item_order += 1
            current["items"].append(
                {"order": item_order, "item_code": f"{current['order']}.{item_order}", "question": line}
            )

    return sections


def load_checklist_sections(*, xlsx_path: Path | None = None) -> list[dict]:
    if xlsx_path is not None:
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {xlsx_path}")
        return parse_checklist_xlsx(xlsx_path)

    if not BUNDLED_CHECKLIST_JSON.exists():
        raise FileNotFoundError(f"Checklist embutido não encontrado: {BUNDLED_CHECKLIST_JSON}")

    data = json.loads(BUNDLED_CHECKLIST_JSON.read_text(encoding="utf-8"))
    return data["sections"]


def seed_checklist_sections(
    sections: list[dict],
    *,
    version_slug: str | None = None,
    activate: bool | None = None,
    replace: bool = False,
) -> tuple[ChecklistVersion, int, int, bool]:
    """Retorna (version, seções, itens, criou)."""
    version_slug = version_slug or date.today().strftime("%Y-%m")
    activate = ChecklistVersion.objects.exists() if activate is None else activate

    if replace:
        ChecklistItem.objects.all().delete()
        ChecklistSection.objects.all().delete()
        ChecklistVersion.objects.all().delete()

    version, created = ChecklistVersion.objects.get_or_create(
        slug=version_slug,
        defaults={"label": f"Anexo IV — {version_slug}", "is_active": activate},
    )
    if not created and activate:
        version.is_active = True
        version.save(update_fields=["is_active"])

    if not replace and version.sections.exists():
        item_count = ChecklistItem.objects.filter(section__version=version).count()
        return version, version.sections.count(), item_count, False

    item_count = 0
    for sec in sections:
        section = ChecklistSection.objects.create(
            version=version,
            order=sec["order"],
            title=sec["title"],
        )
        for item_data in sec["items"]:
            ChecklistItem.objects.create(
                section=section,
                order=item_data["order"],
                item_code=item_data["item_code"],
                question=item_data["question"],
            )
            item_count += 1

    return version, len(sections), item_count, True


def ensure_default_checklist() -> bool:
    """Importa checklist embutido se o banco estiver vazio. Retorna True se importou."""
    from django.db import connection

    tables = set(connection.introspection.table_names())
    if "inspections_checklistversion" not in tables:
        return False

    if ChecklistVersion.objects.exists():
        return False

    sections = load_checklist_sections()
    seed_checklist_sections(sections, version_slug="anexo-iv", activate=True)
    return True
