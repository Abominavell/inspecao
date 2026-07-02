import re
from datetime import date
from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from inspections.models import ChecklistItem, ChecklistSection, ChecklistVersion

DEFAULT_XLSX = Path.home() / "Downloads" / "Anexo IV - Check  List - Diagnóstico de Saúde e Segurança.xlsx"
SECTION_RE = re.compile(r"^(\d+)\.\s+(.+)$")


def parse_checklist(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    start_idx = 0
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any("PREENCHER COM NÚMERO" in c for c in cells):
            start_idx = i + 1
            break
    rows = rows[start_idx:]

    sections: list[dict] = []
    current: dict | None = None
    item_order = 0
    seen_titles: set[str] = set()

    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        line = next((c for c in cells if c), "")

        if line in ("Total", "Total Geral") or line in ("C", "NC", "NA", "Descrição"):
            continue

        sec_match = SECTION_RE.match(line)
        if sec_match and len(line) < 120 and "?" not in line:
            section_num = int(sec_match.group(1))
            title = sec_match.group(2).strip()
            if title.upper().startswith("TOTAL"):
                continue
            full_title = f"{section_num}. {title}"
            if full_title in seen_titles:
                current = next((s for s in sections if s["title"] == full_title), None)
                item_order = len(current["items"]) if current else 0
                continue
            seen_titles.add(full_title)
            current = {"order": section_num, "title": full_title, "items": []}
            sections.append(current)
            item_order = 0
            continue

        if line.upper().startswith("OUTRAS SITUAÇÕES"):
            if line not in seen_titles:
                seen_titles.add(line)
                current = {"order": 21, "title": line, "items": []}
                sections.append(current)
                item_order = 0
            continue

        if current and "?" in line:
            item_order += 1
            current["items"].append(
                {"order": item_order, "item_code": f"{current['order']}.{item_order}", "question": line}
            )

    return sections


class Command(BaseCommand):
    help = "Importa checklist do Anexo IV (Excel) com versionamento"

    def add_arguments(self, parser):
        parser.add_argument("file", nargs="?", type=str, help="Caminho do arquivo .xlsx")
        parser.add_argument(
            "--version",
            type=str,
            default="",
            help="Identificador da versão (ex: 2024-03). Padrão: data atual",
        )
        parser.add_argument(
            "--activate",
            action="store_true",
            help="Define esta versão como ativa para novas inspeções",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            help="Apaga versões antigas e recria do zero (modo legado)",
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options["file"]) if options.get("file") else DEFAULT_XLSX
        if not xlsx_path.exists():
            self.stderr.write(f"Arquivo não encontrado: {xlsx_path}")
            return

        sections = parse_checklist(xlsx_path)
        version_slug = options["version"] or date.today().strftime("%Y-%m")
        activate = options["activate"] or not ChecklistVersion.objects.exists()

        if options["replace"]:
            ChecklistItem.objects.all().delete()
            ChecklistSection.objects.all().delete()
            ChecklistVersion.objects.all().delete()

        version, created = ChecklistVersion.objects.get_or_create(
            slug=version_slug,
            defaults={"label": f"Anexo IV — {version_slug}", "is_active": activate},
        )
        if not created and activate:
            version.is_active = True
            version.save(update_fields=["is_active"])

        if not created and version.sections.exists():
            self.stdout.write(
                self.style.WARNING(
                    f"Versão {version_slug} já possui checklist. Use outro --version ou --replace."
                )
            )
            return

        item_count = 0
        for sec in sections:
            section = ChecklistSection.objects.create(
                version=version,
                order=sec["order"],
                title=sec["title"],
            )
            for item_data in sec["items"]:
                ChecklistItem.objects.create(
                    section=section,
                    order=item_data["order"],
                    item_code=item_data["item_code"],
                    question=item_data["question"],
                )
                item_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Versão {version_slug}: {len(sections)} seções e {item_count} itens de {xlsx_path}"
            )
        )
